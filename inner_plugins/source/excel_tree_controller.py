#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @FileName  :excel_tree_controller.py
# @Time      :2025/4/27 9:07
# @Author    :邓某人,Sheldon
# @Function  :
import json
from pathlib import Path
from typing import List, Union

import openpyxl

mark_column = ["模块", "需求", "负责人"]


class ExcelTreeController:

    def __init__(self, path: Path):
        self.path = path
        self.workbook = None
        self.worksheet = None
        self.data = []
        self.current_row = 0
        self.cache_row_tag = ""
        self.copy_pic_list = set()

    def __enter__(self):
        self.workbook = openpyxl.load_workbook(self.path)
        self.worksheet = self.workbook.active
        self.headers = [cell.value for cell in next(self.worksheet.iter_rows(min_row=1, max_row=1))]
        self.model_header_index = self.headers.index(mark_column[0])
        self.requirement_header_index = self.headers.index(mark_column[1])
        self.pic_header_index = self.headers.index(mark_column[2])
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.workbook.close()

    def __repr__(self):
        return json.dumps(self.__dict__)

    class ModelRow:

        def __init__(self, model: str, requirement: Union[List, None] = None):
            self.model = model
            self.requirement = requirement or []

        def __repr__(self):
            return json.dumps(self.__dict__, default=lambda o: o.__dict__, ensure_ascii=False)

        def append_requirements(self, requirement):
            self.requirement.append(requirement)

        def to_dict(self):
            return {
                "model": self.model,
                "requirement": [req.to_dict() for req in self.requirement]
            }

    class RequirementRow:
        def __init__(self, name: str, pic: Union[List, None] = None):
            self.name = name
            self.pic: List[str] = pic or []

        def __repr__(self):
            return json.dumps(self.__dict__)

        def to_dict(self):
            return {
                "name": self.name,
                "pic": self.pic
            }

    def action(self):
        for row in self.worksheet.iter_rows(min_row=2):
            self.read_row(row)

    def read_row(self, row: tuple['openpyxl.cell.cell.Cell']):
        for idx, cell in enumerate(row):
            if idx == self.model_header_index:
                self.model_handle(cell.value)
            elif idx == self.requirement_header_index:
                self.requirement_handle(cell.value)
            elif idx == self.pic_header_index:
                self.pic_handle(cell.value)

    def model_handle(self, value: Union[str, None]):
        if value is not None and value != self.cache_row_tag:
            self.current_row += 1
            self.cache_row_tag = value
        if len(self.data) < self.current_row:
            self.data.append(self.ModelRow(value))

    def requirement_handle(self, value: Union[str, None]):
        if value is not None:
            row = self.data[-1]
            row.requirement.append(self.RequirementRow(value))

    def pic_handle(self, value: Union[str, None]):
        row = self.data[-1]
        if value is not None:
            row.requirement[-1].pic.append(value)
            self.copy_pic_list.add(value)
        else:
            row.requirement[-1].pic.append(row.requirement[-2].pic[0])

    def get_data(self):
        _data = self.de_duplicates(self.data)
        return [model.to_dict() for model in self.change_to_obj(_data)]

    @staticmethod
    def de_duplicates(data: List):
        cache_model = {}
        for index, item in enumerate(data):
            _item: ExcelTreeController.ModelRow = item
            _model = cache_model.get(_item.model, None)
            if _model is None:
                cache_model[_item.model] = {}
            for requirement in _item.requirement:
                _requirement = cache_model[_item.model].get(requirement.name, None)
                if _requirement is None:
                    cache_model[_item.model][requirement.name] = set()
                for pic in requirement.pic:
                    cache_model[_item.model][requirement.name].add(pic)
        return cache_model

    @staticmethod
    def change_to_obj(data: dict):
        cache_list = []
        for key, value in data.items():
            _model = ExcelTreeController.ModelRow(key)
            for re_name, re_pic_set in value.items():
                _requirement = ExcelTreeController.RequirementRow(re_name, list(re_pic_set))
                _model.append_requirements(_requirement)
            cache_list.append(_model)
        return cache_list


def get_excel_tree_dict(path: Path):
    with ExcelTreeController(path) as excel_tree:
        excel_tree.action()
        return excel_tree.get_data(), list(excel_tree.copy_pic_list)


if __name__ == '__main__':
    print(get_excel_tree_dict(Path('/Users/sheldon/Documents/Sheldon/ttttt/work_test.xlsx')))
