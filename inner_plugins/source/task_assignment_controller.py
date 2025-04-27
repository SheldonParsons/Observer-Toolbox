#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @FileName  :task_assignment_controller.py
# @Time      :2025/4/27 9:07
# @Author    :邓某人
# @Function  :
from openpyxl.reader.excel import load_workbook


def read_specific_columns(file_path: str, sheet_index: int = 0):
    try:
        wb = load_workbook(file_path)
        # sheet = wb.active
        sheet = wb.worksheets[sheet_index]
    except FileNotFoundError:
        print(f"错误：未找到文件 {file_path}。")
        raise
    except Exception as e:
        print(f"读取文件时出现未知错误: {e}")
        raise
    header = [cell.value for cell in sheet[1]]
    module_col = header.index('所属模块') + 1
    requirement_col = header.index('需求') + 1
    person_col = header.index('负责人') + 1
    if module_col is None or requirement_col is None or person_col is None:
        raise ValueError("表头中缺少必要的列：所属模块、需求或负责人")
    merged_cell_map = {}  # 缓存合并单元格的数据
    for merge_range in sheet.merged_cells.ranges:  # 遍历所有合并单元格
        start_row = merge_range.min_row
        end_row = merge_range.max_row
        module_value = sheet.cell(start_row, merge_range.min_col).value
        for row in range(start_row, end_row + 1):
            merged_cell_map[row] = module_value

    data = []
    for row_num in range(2, sheet.max_row + 1):
        #   优先从缓存里面拿数据
        module_value = merged_cell_map.get(row_num, sheet.cell(row_num, module_col).value)
        requirement = sheet.cell(row_num, requirement_col).value
        person = sheet.cell(row_num, person_col).value

        if requirement is None and person is None:
            continue

        module = module_value if module_value is not None else 'empty_model'
        data.append({
            'module': module,
            'requirement': requirement,
            'person': person
        })
    module_list = set()
    owner_list = set()
    requirement_details = {}

    for item in data:
        module = item['module']
        requirement = item['requirement']
        person = item['person']

        if module:
            module_list.add(module)
        if person:
            owner_list.add(person)
        if module not in requirement_details:
            requirement_details[module] = []
        if requirement:
            requirement_details[module].append(str(requirement).strip())
    result = [
        {'module_list': list(module_list)},
        {'owner_list': list(owner_list)},
        {'requirement_details': requirement_details}
    ]
    # pprint(result)
    return result


if __name__ == '__main__':
    read_specific_columns(file_path='测试任务分配表.xlsx')
